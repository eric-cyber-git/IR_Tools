import re
import pandas as pd
import shutil
import os
import json
import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

# === STAGE 1: CASE MANAGEMENT ===
def get_next_case_number(case_root_path):
    case_numbers = [int(num) for num in os.listdir(case_root_path)]
    return max(case_numbers) + 1

def create_case_folder_structure(case_root_path, case_number):
    case_path = os.path.join(case_root_path, str(case_number))
    os.mkdir(case_path)
    
    baseline_path = os.path.join(case_path, "Baseline")
    target_path = os.path.join(case_path, "Target")
    output_path = os.path.join(case_path, "Output")
    os.mkdir(output_path)

    return {
        "case": case_path,
        "baseline": baseline_path,
        "target": target_path,
        "output": output_path
    }

# === STAGE 2: FILE PREP ===
def move_case_txt_files(source_folder, dest_folder):
    for file in os.listdir(source_folder):
        if file.endswith('.txt'):
            shutil.move(os.path.join(source_folder, file), dest_folder)

def move_target_and_baseline_files(target_src, target_dest, baseline_src, baseline_dest):
    shutil.move(target_src, target_dest)
    shutil.copytree(baseline_src, baseline_dest)

def rename_files_with_prefix(path, prefix):
    renamed_files = []
    for file in os.listdir(path):
        new_name = f"{prefix}_{file}"
        src = os.path.join(path, file)
        dst = os.path.join(path, new_name)
        os.rename(src, dst)
        renamed_files.append(dst)
    return renamed_files

def move_files_to_output(src_folder, dest_folder):
    for file in os.listdir(src_folder):
        shutil.move(os.path.join(src_folder, file), dest_folder)

# === STAGE 3: CSV MERGE AND CLEAN ===
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def merge_csvs_to_excel(output_folder):
    csv_file_path = os.listdir(output_folder)
    csv_data_dict = []

    for file in csv_file_path:
        if not file.endswith(".csv"):
            continue
        formatted_name = file.split(".")[0]
        print(formatted_name)
        df = pd.read_csv(os.path.join(output_folder, file))
        csv_data_dict.append({"Sheet_Name": formatted_name, "Data_Frame": df})

    with pd.ExcelWriter(os.path.join(output_folder, "unified_book.xlsx"), engine="xlsxwriter") as writer:
        for entry in csv_data_dict:
            df_clean = entry["Data_Frame"].map(lambda x: ILLEGAL_CHARACTERS_RE.sub('---', x) if isinstance(x, str) else x)
            df_clean.to_excel(writer, sheet_name=entry["Sheet_Name"], index=False)

# === STAGE 4: CONDITIONAL FORMATTING ===
def apply_conditional_formatting(config_path, output_folder, case_number):
    xl_path = os.path.join(output_folder, "unified_book.xlsx")
    xl_workbook = openpyxl.load_workbook(xl_path)

    with open(config_path) as f:
        config_data = json.load(f)

    yellowfill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for entry in config_data:
        try:
            sheet_name = entry["Sheet_name"]
            cell_range = entry["Cell_Range"]
            formula = entry["Formula"]

            sheet = xl_workbook[sheet_name]
            sheet.conditional_formatting.add(cell_range, FormulaRule(formula=[formula], fill=yellowfill))
        except Exception as e:
            print(f"Formatting failed for sheet {entry['Sheet_name']}: {e}")

    final_name = f"01_Formatted_unified_book_{str(case_number)}.xlsx"
    xl_workbook.save(os.path.join(output_folder, final_name))
    os.system(f'start EXCEL.EXE "{os.path.join(output_folder, final_name)}"')

# === MAIN EXECUTION ===
def main():
    WinIR_Config_Folder = os.getenv('WinIR_Config_Folder')
    WinIR_Case_folder = os.getenv('WinIR_Case_folder')
    case_data = os.getenv('WinIR_Case_Data')
    baseline_backup = os.getenv('WinIR_Baseline_Backup')

    case_number = get_next_case_number(WinIR_Case_folder)
    paths = create_case_folder_structure(WinIR_Case_folder, case_number)

    move_case_txt_files(case_data, paths["case"])
    move_target_and_baseline_files(
        case_data,
        paths["target"],
        baseline_backup,
        paths["baseline"]
    )

    rename_files_with_prefix(paths["baseline"], "baseline")
    rename_files_with_prefix(paths["target"], "target")

    move_files_to_output(paths["baseline"], paths["output"])
    move_files_to_output(paths["target"], paths["output"])

    merge_csvs_to_excel(paths["output"])
    apply_conditional_formatting(
        os.path.join(WinIR_Config_Folder, "CF_config.json"),
        paths["output"],
        case_number
    )

if __name__ == "__main__":
    main()
